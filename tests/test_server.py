from __future__ import annotations

import base64
import io
import json
import os
import tempfile
import threading
import unittest
from datetime import date
from http.cookiejar import CookieJar
from http.server import ThreadingHTTPServer
from pathlib import Path
from urllib.error import HTTPError
from urllib.request import HTTPCookieProcessor, Request, build_opener

from openpyxl import Workbook, load_workbook
from pypdf import PdfWriter


TEST_ROOT = tempfile.TemporaryDirectory(prefix="neurum-renal-test-")
os.environ["RENAL_DATA_DIR"] = str(Path(TEST_ROOT.name) / "data")
os.environ["RENAL_BACKUP_DIR"] = str(Path(TEST_ROOT.name) / "backups")
os.environ["RENAL_ADMIN_EMAIL"] = "admin@test.local"
os.environ["RENAL_ADMIN_PASSWORD"] = "TestAdmin12345!"
os.environ["RENAL_COOKIE_SECURE"] = "false"
os.environ["RENAL_QUIET"] = "true"

from database import connect, initialize  # noqa: E402
from cac_transformer import (  # noqa: E402
    SupportEvidence,
    choose_lab,
    choose_vitals,
    extract_support_vitals,
    process_cac_job,
)
from server import RenalRequestHandler  # noqa: E402


class ClinicalImportApiTest(unittest.TestCase):
    @classmethod
    def setUpClass(cls) -> None:
        initialize()
        cls.server = ThreadingHTTPServer(("127.0.0.1", 0), RenalRequestHandler)
        cls.thread = threading.Thread(target=cls.server.serve_forever, daemon=True)
        cls.thread.start()
        cls.base_url = f"http://127.0.0.1:{cls.server.server_port}"
        cls.opener = build_opener(HTTPCookieProcessor(CookieJar()))
        login = cls.request(
            "/api/auth/login",
            "POST",
            {"email": "admin@test.local", "password": "TestAdmin12345!"},
            csrf=False,
        )
        cls.csrf = login["csrf_token"]

    @classmethod
    def tearDownClass(cls) -> None:
        cls.server.shutdown()
        cls.server.server_close()
        cls.thread.join(timeout=5)
        TEST_ROOT.cleanup()

    @classmethod
    def request(cls, path: str, method: str = "GET", payload: dict | None = None, csrf: bool = True) -> dict:
        data = json.dumps(payload).encode("utf-8") if payload is not None else None
        headers = {"Accept": "application/json"}
        if data is not None:
            headers["Content-Type"] = "application/json"
        if csrf and method != "GET":
            headers["X-CSRF-Token"] = cls.csrf
        try:
            response = cls.opener.open(
                Request(
                    cls.base_url + path,
                    data=data,
                    headers=headers,
                    method=method,
                ),
                timeout=60,
            )
        except HTTPError as exc:
            detail = exc.read().decode("utf-8", errors="replace")
            raise AssertionError(
                f"{method} {path} respondió HTTP {exc.code}: {detail}"
            ) from exc
        return json.loads(response.read().decode("utf-8"))

    def test_cohort_and_lab_batches_are_persisted_as_one_import(self) -> None:
        import_id = "test-import-001"
        cohort = self.request(
            "/api/clinical/sync",
            "POST",
            {
                "mode": "cohort",
                "patients": [{"id": "1001", "type": "CC", "name": "Paciente Uno", "activeInCohort": True}],
                "labs": [],
                "file_name": "cohorte.xlsx",
                "import_id": import_id,
                "import_start": True,
                "import_final": False,
            },
        )
        self.assertEqual(cohort["patients"], 1)

        first_batch = self.request(
            "/api/clinical/sync",
            "POST",
            {
                "mode": "labs",
                "labs": [
                    self.lab("creatinine-1", "1001", "creatinine", "2026-01-10", 1.1),
                    self.lab("hba1c-1", "1001", "hba1c", "2026-02-10", 7.2),
                    self.lab("unknown-1", "9999", "creatinine", "2026-01-10", 1.0),
                ],
                "file_name": "cohorte.xlsx",
                "import_id": import_id,
                "import_final": False,
            },
        )
        self.assertEqual(first_batch["labs_added"], 2)
        self.assertEqual(first_batch["labs_skipped"], 1)

        final_batch = self.request(
            "/api/clinical/sync",
            "POST",
            {
                "mode": "labs",
                "labs": [
                    self.lab("creatinine-1", "1001", "creatinine", "2026-01-10", 1.1),
                    self.lab("acr-1", "1001", "acr", "2026-03-10", 24),
                ],
                "file_name": "cohorte.xlsx",
                "import_id": import_id,
                "import_final": True,
            },
        )
        self.assertEqual(final_batch["labs_added"], 1)

        snapshot = self.request("/api/clinical/snapshot")
        self.assertEqual(len(snapshot["patients"]), 1)
        self.assertEqual(len(snapshot["labs"]), 3)

        imports = self.request("/api/imports")["imports"]
        imported = next(item for item in imports if item["id"] == import_id)
        self.assertEqual(imported["record_count"], 1)
        self.assertEqual(imported["lab_count"], 3)

        with connect() as conn:
            patient = conn.execute("SELECT doc_type,document,active FROM patients WHERE id='1001'").fetchone()
        self.assertEqual(dict(patient), {"doc_type": "CC", "document": "1001", "active": 1})

    def test_task_selected_from_cohort_moves_through_management_and_resolution(self) -> None:
        created = self.request(
            "/api/tasks/from-exam",
            "POST",
            {
                "patient_id": "1001",
                "exam_type": "creatinine",
                "managed": True,
                "due_date": "2026-07-30",
                "priority": "alta",
            },
        )["task"]
        self.assertEqual(created["status"], "en_gestion")
        self.assertEqual(created["patient_name"], "Paciente Uno")
        self.assertEqual(created["assignee_name"], "Administrador Neurum")

        listed = self.request("/api/tasks")["tasks"]
        self.assertTrue(any(task["id"] == created["id"] for task in listed))

        snapshot = self.request("/api/clinical/snapshot")
        managed = next(
            task
            for task in snapshot["managed_exams"]
            if task["patient_id"] == "1001" and task["exam_type"] == "creatinine"
        )
        self.assertEqual(managed["status"], "en_gestion")

        resolved = self.request(
            f"/api/tasks/{created['id']}",
            "PATCH",
            {"status": "completada", "note": "Paciente citado y examen realizado."},
        )["task"]
        self.assertEqual(resolved["status"], "completada")
        self.assertEqual(resolved["closer_name"], "Administrador Neurum")
        self.assertIsNotNone(resolved["closed_at"])

        history = self.request(f"/api/tasks/{created['id']}/history")["events"]
        self.assertEqual(history[0]["to_status"], "completada")
        self.assertEqual(history[0]["note"], "Paciente citado y examen realizado.")

        reopened = self.request(
            "/api/tasks/from-exam",
            "POST",
            {"patient_id": "1001", "exam_type": "creatinine", "managed": False},
        )["task"]
        self.assertEqual(reopened["status"], "pendiente")
        self.assertIsNone(reopened["closed_at"])

    def test_cac_job_generates_audited_xlsx_and_ready_txt(self) -> None:
        created = self.request(
            "/api/cac/jobs",
            "POST",
            {
                "settings": {
                    "cutoff_date": "2026-06-30",
                    "eapb_code": "EPS040",
                    "ips_code": "050011399101",
                    "affiliation_date": "2020-01-01",
                    "ethnicity_code": "6",
                    "population_group": "61",
                    "territorial_entity": True,
                    "htn_cost_default": 0,
                    "dm_cost_default": 0,
                    "total_cost_default": 0,
                }
            },
        )
        job_id = created["job"]["id"]

        workbook = Workbook()
        sheet = workbook.active
        headers = [
            "numero_documento",
            "fechaCreacion",
            "nombres_paciente",
            "apellidos_paciente",
            "tipo_documento",
            "sexo",
            "regimen",
            "municipio_residencia",
            "fecha_nacimiento",
            "tipo_diabetes",
            "fecha_diagnostico_diabetes",
            "hta",
            "peso",
            "talla",
            "sistolica",
            "diastolica",
            "creatinine",
            "hba1c",
            "rac",
            "tfg",
            "fecha_paraclinicos",
            "nefropatia_diabetica",
            "trr",
            "transplant",
        ]
        sheet.append(headers)
        sheet.append(
            [
                "1001",
                "2026-06-20",
                "Paciente Uno",
                "Prueba Control",
                "Cédula ciudadanía",
                "Femenino",
                "Subsidiado",
                "Medellin",
                "1980-01-01",
                "Tipo 2",
                "2010-01-01",
                "no",
                65,
                1.6,
                120,
                75,
                0.8,
                7.1,
                20,
                100,
                "2026-05-15",
                "no",
                "No",
                "no",
            ]
        )
        sheet.append(
            [
                "1002",
                "2026-06-20",
                "Paciente Dos",
                "Prueba Sin TFG",
                "Cédula ciudadanía",
                "Femenino",
                "Subsidiado",
                "Medellin",
                "1985-01-01",
                "Tipo 2",
                "2012-01-01",
                "no",
                62,
                1.58,
                118,
                72,
                0.9,
                6.9,
                18,
                None,
                "2026-05-15",
                "no",
                "No",
                "no",
            ]
        )
        xlsx_buffer = io.BytesIO()
        workbook.save(xlsx_buffer)
        self.request(
            f"/api/cac/jobs/{job_id}/consultations",
            "POST",
            {
                "file_name": "atenciones.xlsx",
                "data_base64": base64.b64encode(xlsx_buffer.getvalue()).decode("ascii"),
            },
        )

        pdf_writer = PdfWriter()
        pdf_writer.add_blank_page(width=612, height=792)
        pdf_buffer = io.BytesIO()
        pdf_writer.write(pdf_buffer)
        second_pdf_writer = PdfWriter()
        second_pdf_writer.add_blank_page(width=600, height=780)
        second_pdf_buffer = io.BytesIO()
        second_pdf_writer.write(second_pdf_buffer)
        self.request(
            f"/api/cac/jobs/{job_id}/supports",
            "POST",
            {
                "file_name": "CC1001_MEDICINA INTERNA_2026-06-20.pdf",
                "data_base64": base64.b64encode(pdf_buffer.getvalue()).decode("ascii"),
            },
        )
        self.request(
            f"/api/cac/jobs/{job_id}/supports",
            "POST",
            {
                "file_name": "CC1002_MEDICINA INTERNA_2026-06-20.pdf",
                "data_base64": base64.b64encode(second_pdf_buffer.getvalue()).decode(
                    "ascii"
                ),
            },
        )

        processed = self.request(f"/api/cac/jobs/{job_id}/process", "POST", {})
        self.assertEqual(processed["summary"]["patients"], 2)
        self.assertEqual(processed["summary"]["matched_supports"], 2)
        self.assertEqual(processed["summary"]["ready"], 1)
        self.assertEqual(processed["summary"]["review"], 1)
        self.assertTrue(processed["summary"]["txt_name"].endswith("_ERC.txt"))

        detail = self.request(f"/api/cac/jobs/{job_id}")
        self.assertEqual(detail["job"]["status"], "completado")
        ready_patient = next(
            patient
            for patient in detail["preview"]["patients"]
            if patient["document"] == "1001"
        )
        missing_tfg_patient = next(
            patient
            for patient in detail["preview"]["patients"]
            if patient["document"] == "1002"
        )
        self.assertEqual(ready_patient["status"], "LISTO")
        self.assertEqual(ready_patient["key_values"]["tfg"], "100")
        self.assertEqual(ready_patient["key_values"]["systolic"], "120")
        self.assertEqual(ready_patient["key_values"]["diastolic"], "75")
        self.assertEqual(missing_tfg_patient["status"], "REVISAR")
        self.assertEqual(missing_tfg_patient["key_values"]["tfg"], 999)
        self.assertEqual(missing_tfg_patient["key_values"]["erc_stage"], 99)
        self.assertTrue(
            any(
                "TFG no registrada en la atención" in error
                for error in missing_tfg_patient["errors"]
            )
        )
        self.assertFalse(
            any(
                "calculada" in evidence.lower()
                for evidence in missing_tfg_patient["evidence"]
            )
        )

        with connect() as conn:
            job = conn.execute(
                "SELECT output_xlsx_path,output_txt_path FROM cac_jobs WHERE id=?",
                (job_id,),
            ).fetchone()
        self.assertTrue(Path(job["output_xlsx_path"]).is_file())
        self.assertTrue(Path(job["output_txt_path"]).is_file())

    def test_cac_consolidates_encounters_and_keeps_latest_lab_dates(self) -> None:
        with tempfile.TemporaryDirectory(prefix="cac-consolidation-") as directory:
            root = Path(directory)
            source = root / "atenciones.xlsx"
            output = root / "output"
            workbook = Workbook()
            sheet = workbook.active
            headers = [
                "numero_documento",
                "fechaCreacion",
                "nombres_paciente",
                "apellidos_paciente",
                "tipo_documento",
                "sexo",
                "regimen",
                "municipio_residencia",
                "fecha_nacimiento",
                "tipo_diabetes",
                "fecha_diagnostico_diabetes",
                "hta",
                "peso",
                "talla",
                "sistolica",
                "diastolica",
                "creatinine",
                "hba1c",
                "rac",
                "tfg",
                "fecha_paraclinicos",
                "paraclinicos",
                "nefropatia_diabetica",
                "trr",
                "transplant",
            ]
            sheet.append(headers)
            sheet.append(
                [
                    "777",
                    "2026-01-20",
                    "Paciente",
                    "Consolidado",
                    "Permiso de Proteccion Termporal",
                    "Femenino",
                    "Subsidiado",
                    "Medellin",
                    "1980-01-01",
                    "Tipo 2",
                    "2010-01-01",
                    "no",
                    64,
                    1.6,
                    120,
                    75,
                    1.2,
                    None,
                    None,
                    88,
                    "2026-01-15",
                    None,
                    "no",
                    "no",
                    "no",
                ]
            )
            sheet.append(
                [
                    "777",
                    "2026-03-05",
                    "Paciente",
                    "Consolidado",
                    None,
                    None,
                    None,
                    None,
                    None,
                    None,
                    None,
                    None,
                    None,
                    None,
                    None,
                    None,
                    None,
                    None,
                    None,
                    None,
                    None,
                    "* 02.03.2026 HbA1c 7.3",
                    None,
                    None,
                    None,
                ]
            )
            sheet.append(
                [
                    "777",
                    "2026-04-10",
                    "Paciente",
                    "Consolidado",
                    None,
                    None,
                    None,
                    None,
                    None,
                    None,
                    None,
                    None,
                    None,
                    None,
                    None,
                    None,
                    None,
                    None,
                    22,
                    None,
                    "2026-04-08",
                    None,
                    None,
                    None,
                    None,
                ]
            )
            workbook.save(source)

            preview = process_cac_job(
                source,
                [],
                output,
                {
                    "cutoff_date": "2026-06-30",
                    "eapb_code": "EPS040",
                    "ips_code": "050011399101",
                    "affiliation_date": "2020-01-01",
                    "ethnicity_code": "6",
                    "population_group": "61",
                    "territorial_entity": True,
                    "htn_cost_default": 0,
                    "dm_cost_default": 0,
                    "total_cost_default": 0,
                },
            )

            self.assertEqual(preview["summary"]["patients"], 1)
            self.assertEqual(preview["summary"]["source_records"], 3)
            key = preview["patients"][0]["key_values"]
            self.assertEqual(key["creatinine"], "1.2")
            self.assertEqual(key["creatinine_date"], "2026-01-15")
            self.assertEqual(key["hba1c"], "7.3")
            self.assertEqual(key["hba1c_date"], "2026-03-02")
            self.assertEqual(key["rac"], "22")
            self.assertEqual(key["rac_date"], "2026-04-08")

            generated = load_workbook(output / "malla_cac_erc_validada.xlsx")
            malla = generated["Malla_CAC"]
            header_positions = {
                cell.value: cell.column
                for cell in malla[2]
                if cell.value
            }
            self.assertEqual(
                malla.cell(3, header_positions["VAR5_TipoIdentificacion"]).value,
                "PT",
            )
            self.assertEqual(malla.max_row, 3)

    def test_cac_prefers_the_newest_pdf_lab_over_an_older_attention(self) -> None:
        records = [
            {
                "hba1c": 7.1,
                "fecha_paraclinicos": "2026-02-10",
                "_encounter_date": date(2026, 2, 15),
                "_narrative_labs": [],
            }
        ]
        support = SupportEvidence(
            file_name="CC777_ENDOCRINOLOGIA_2026-04-20.pdf",
            document="777",
            doc_type="CC",
            attention_date=date(2026, 4, 20),
            specialty="ENDOCRINOLOGIA",
            text="",
            normalized_text="",
            phone="",
            ips_code="",
            labs=[
                {
                    "type": "hba1c",
                    "date": date(2026, 4, 18),
                    "value": 8.2,
                    "source": "CC777_ENDOCRINOLOGIA_2026-04-20.pdf",
                    "source_kind": "pdf_structured",
                    "source_rank": 4,
                    "encounter_date": date(2026, 4, 20),
                }
            ],
        )
        warnings: list[str] = []

        selected = choose_lab(records, [support], "hba1c", "hba1c", warnings)

        self.assertIsNotNone(selected)
        self.assertEqual(selected["value"], 8.2)
        self.assertEqual(selected["date"], date(2026, 4, 18))
        self.assertTrue(any("PDF actualiza" in warning for warning in warnings))

    def test_cac_uses_each_latest_available_vital_without_blank_overwrite(self) -> None:
        records = [
            {
                "peso": 80,
                "talla": 170,
                "sistolica": 120,
                "diastolica": 70,
                "_encounter_date": date(2026, 1, 10),
            },
            {
                "peso": "",
                "talla": "",
                "sistolica": 130,
                "diastolica": 80,
                "_encounter_date": date(2026, 4, 10),
            },
        ]
        older_support = SupportEvidence(
            file_name="CC777_MEDICINA_INTERNA_2026-02-20.pdf",
            document="777",
            doc_type="CC",
            attention_date=date(2026, 2, 20),
            specialty="MEDICINA INTERNA",
            text="",
            normalized_text="",
            phone="",
            ips_code="",
            labs=[],
            vitals=[
                {
                    "type": "weight",
                    "value": 75,
                    "date": date(2026, 2, 20),
                    "source": "CC777_MEDICINA_INTERNA_2026-02-20.pdf",
                    "source_kind": "pdf_vital",
                    "source_rank": 2,
                },
                {
                    "type": "height",
                    "value": 169,
                    "date": date(2026, 2, 20),
                    "source": "CC777_MEDICINA_INTERNA_2026-02-20.pdf",
                    "source_kind": "pdf_vital",
                    "source_rank": 2,
                },
            ],
        )
        newest_blank_support = SupportEvidence(
            file_name="CC777_ENDOCRINOLOGIA_2026-05-20.pdf",
            document="777",
            doc_type="CC",
            attention_date=date(2026, 5, 20),
            specialty="ENDOCRINOLOGIA",
            text="",
            normalized_text="",
            phone="",
            ips_code="",
            labs=[],
            vitals=[],
        )

        selected = choose_vitals(records, [older_support, newest_blank_support])

        self.assertEqual(selected["weight"]["value"], 75)
        self.assertEqual(selected["weight"]["date"], date(2026, 2, 20))
        self.assertEqual(selected["height"]["value"], 169)
        self.assertEqual(selected["systolic"]["value"], 130)
        self.assertEqual(selected["systolic"]["date"], date(2026, 4, 10))
        self.assertEqual(selected["diastolic"]["value"], 80)

    def test_cac_extracts_interleaved_pdf_vital_table(self) -> None:
        layout = """
  Signos vitales
  Presión Arterial (Sistólica)                  Frecuencia respiratoria (RPM)                Saturación de oxígeno (%)
  114                                          16                                           Peso
  Presión Arterial (Diastólica)                Temperatura (°C)                             76.7
  76                                           36                                           Talla
  Frecuencia cardiaca (LPM)                    Perímetro abdominal                          156
  Examen físico
"""

        vitals = extract_support_vitals(
            [layout],
            "CC777_ENDOCRINOLOGIA_2026-01-20.pdf",
            date(2026, 1, 20),
        )
        values = {vital["type"]: vital["value"] for vital in vitals}

        self.assertEqual(values["systolic"], 114)
        self.assertEqual(values["diastolic"], 76)
        self.assertEqual(str(values["weight"]), "76.7")
        self.assertEqual(values["height"], 156)

    @staticmethod
    def lab(key: str, patient_id: str, lab_type: str, date: str, value: float) -> dict:
        return {
            "key": key,
            "patientId": patient_id,
            "type": lab_type,
            "date": date,
            "value": value,
            "source": "Laboratorio de prueba",
        }


if __name__ == "__main__":
    unittest.main()
