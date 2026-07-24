from __future__ import annotations

import json
import math
import re
import unicodedata
from collections import Counter, defaultdict
from dataclasses import dataclass, field
from datetime import date, datetime
from decimal import Decimal, InvalidOperation, ROUND_DOWN
from pathlib import Path
from typing import Any

from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from pypdf import PdfReader


BASE_DIR = Path(__file__).resolve().parent
TEMPLATE_PATH = BASE_DIR / "resources" / "malla_cac_erc_2026.xlsx"
DIVIPOLA_PATH = BASE_DIR / "resources" / "divipola_map.json"
REPORT_START = date(2025, 7, 1)
REPORT_END = date(2026, 6, 30)

DOCUMENT_TYPES = {
    "registro civil": "RC",
    "tarjeta identidad": "TI",
    "tarjeta de identidad": "TI",
    "cedula ciudadania": "CC",
    "cedula de ciudadania": "CC",
    "cedula extranjeria": "CE",
    "cedula de extranjeria": "CE",
    "pasaporte": "PA",
    "menor sin identificacion": "MS",
    "adulto sin identificacion": "AS",
    "carne diplomatico": "CD",
    "salvoconducto": "SC",
    "permiso especial": "PE",
    "permiso de proteccion temporal": "PT",
    "permiso de proteccion termporal": "PT",
    "permiso proteccion termporal": "PT",
    "sin identificacion": "SI",
    "documento extranjero": "DE",
    "certificado nacido vivo": "CN",
    "salvo conducto": "SC",
    "cc": "CC",
    "ti": "TI",
    "pt": "PT",
    "ce": "CE",
    "pa": "PA",
    "rc": "RC",
}
SEX_CODES = {"femenino": "F", "mujer": "F", "f": "F", "masculino": "M", "hombre": "M", "m": "M"}
REGIMEN_CODES = {
    "contributivo": "C",
    "subsidiado": "S",
    "regimen de excepcion": "P",
    "excepcion": "P",
    "especial": "E",
    "no asegurado": "N",
    "planes voluntarios": "V",
    "ppl": "I",
}
DM_CODES = {"tipo 1": 1, "tipo1": 1, "tipo 2": 3, "tipo2": 3, "otro": 4, "otros": 4}
IECA_NAMES = (
    "enalapril",
    "lisinopril",
    "captopril",
    "ramipril",
    "perindopril",
    "benazepril",
)
ARA2_NAMES = (
    "losartan",
    "valsartan",
    "irbesartan",
    "candesartan",
    "telmisartan",
    "olmesartan",
    "azilsartan",
)
LAB_LABELS = {
    "creatinina": "creatinine",
    "hba1c": "hba1c",
    "albuminuria": "albuminuria",
    "rac": "rac",
    "colesterol ldl": "ldl",
    "colesterol hdl": "hdl",
    "colesterol total": "total_cholesterol",
}
NARRATIVE_DATE_PATTERN = re.compile(
    r"(?<!\d)(?P<day>\d{1,2})\s*[./-]\s*(?P<month>\d{1,2})"
    r"\s*[./-]\s*(?P<year>\d{2,4})(?!\d)"
)
NARRATIVE_LAB_PATTERNS = {
    "creatinine": re.compile(
        r"(?<![A-Z0-9])(?:CREATININA|CREATININE|CR)(?![A-Z])"
        r"\s*(?:[:=]|DE)?\s*([0-9]+(?:[.,][0-9]+)?)",
        re.I,
    ),
    "hba1c": re.compile(
        r"(?<![A-Z0-9])(?:HBA1C|HB\s*A1C|HEMOGLOBINA\s+GLICOSILADA)"
        r"(?![A-Z0-9])\s*(?:[:=]|DE)?\s*([0-9]+(?:[.,][0-9]+)?)",
        re.I,
    ),
    "albuminuria": re.compile(
        r"(?<![A-Z0-9])(?:MICROALBUMINURIA|ALBUMINURIA)(?![A-Z0-9])"
        r"\s*(?:[:=]|DE)?\s*([0-9]+(?:[.,][0-9]+)?)",
        re.I,
    ),
    "rac": re.compile(
        r"(?<![A-Z0-9])RAC(?![A-Z0-9])"
        r"\s*(?:[:=]|DE)?\s*([0-9]+(?:[.,][0-9]+)?)",
        re.I,
    ),
    "ldl": re.compile(
        r"(?<![A-Z0-9])(?:C[-\s]?)?LDL(?![A-Z0-9])"
        r"\s*(?:[:=]|DE)?\s*([0-9]+(?:[.,][0-9]+)?)",
        re.I,
    ),
    "hdl": re.compile(
        r"(?<![A-Z0-9])(?:C[-\s]?)?HDL(?![A-Z0-9])"
        r"\s*(?:[:=]|DE)?\s*([0-9]+(?:[.,][0-9]+)?)",
        re.I,
    ),
    "total_cholesterol": re.compile(
        r"(?<![A-Z0-9])(?:COLESTEROL\s+TOTAL|CT)(?![A-Z0-9])"
        r"\s*(?:[:=]|DE)?\s*([0-9]+(?:[.,][0-9]+)?)",
        re.I,
    ),
}
NARRATIVE_LAB_RANGES = {
    "creatinine": (Decimal("0.01"), Decimal("50")),
    "hba1c": (Decimal("1"), Decimal("30")),
    "albuminuria": (Decimal("0"), Decimal("100000")),
    "rac": (Decimal("0"), Decimal("100000")),
    "ldl": (Decimal("1"), Decimal("2000")),
    "hdl": (Decimal("1"), Decimal("1000")),
    "total_cholesterol": (Decimal("1"), Decimal("3000")),
}
VITAL_FIELDS = {
    "weight": {
        "attention_field": "peso",
        "label": re.compile(r"Peso", re.I),
        "minimum": Decimal("1"),
        "maximum": Decimal("400"),
    },
    "height": {
        "attention_field": "talla",
        "label": re.compile(r"Talla", re.I),
        "minimum": Decimal("30"),
        "maximum": Decimal("250"),
    },
    "systolic": {
        "attention_field": "sistolica",
        "label": re.compile(r"Presi[oó]n\s+Arterial\s+\(Sist[oó]lica\)", re.I),
        "minimum": Decimal("50"),
        "maximum": Decimal("300"),
    },
    "diastolic": {
        "attention_field": "diastolica",
        "label": re.compile(r"Presi[oó]n\s+Arterial\s+\(Diast[oó]lica\)", re.I),
        "minimum": Decimal("30"),
        "maximum": Decimal("200"),
    },
}


@dataclass
class SupportEvidence:
    file_name: str
    document: str
    doc_type: str
    attention_date: date | None
    specialty: str
    text: str
    normalized_text: str
    phone: str
    ips_code: str
    labs: list[dict[str, Any]]
    vitals: list[dict[str, Any]] = field(default_factory=list)


def normalize(value: Any) -> str:
    text = unicodedata.normalize("NFKD", str(value or ""))
    text = "".join(char for char in text if not unicodedata.combining(char))
    return re.sub(r"\s+", " ", text).strip().lower()


def normalize_key(value: Any) -> str:
    return re.sub(r"[^a-z0-9]+", " ", normalize(value)).strip()


def ascii_text(value: Any) -> str:
    text = unicodedata.normalize("NFKD", str(value or ""))
    text = "".join(char for char in text if not unicodedata.combining(char))
    text = re.sub(r"[^A-Za-z0-9 .,_:/()+%-]", "", text)
    return re.sub(r"\s+", " ", text).strip()


def clean_document(value: Any) -> str:
    text = str(value or "").strip()
    if text.endswith(".0") and text[:-2].isdigit():
        text = text[:-2]
    return re.sub(r"\s+", "", text)


def meaningful(value: Any) -> bool:
    if value is None:
        return False
    if isinstance(value, float) and math.isnan(value):
        return False
    return str(value).strip() not in {"", "-", "None", "nan", "NaN"}


def parse_date(value: Any) -> date | None:
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    text = str(value or "").strip()
    if not text:
        return None
    text = text.split(" ")[0]
    for pattern in (
        "%Y-%m-%d",
        "%d-%m-%Y",
        "%d/%m/%Y",
        "%d/%m/%y",
        "%d.%m.%Y",
        "%d.%m.%y",
        "%Y/%m/%d",
    ):
        try:
            return datetime.strptime(text, pattern).date()
        except ValueError:
            continue
    return None


def date_text(value: Any, fallback: str = "") -> str:
    parsed = parse_date(value)
    return parsed.isoformat() if parsed else fallback


def decimal_value(value: Any) -> Decimal | None:
    if not meaningful(value):
        return None
    text = str(value).strip().replace(",", ".")
    match = re.search(r"-?\d+(?:\.\d+)?", text)
    if not match:
        return None
    try:
        return Decimal(match.group(0))
    except InvalidOperation:
        return None


def truncated(value: Any, places: int = 2) -> str:
    number = decimal_value(value)
    if number is None:
        return ""
    quantum = Decimal(1).scaleb(-places)
    result = number.quantize(quantum, rounding=ROUND_DOWN)
    rendered = f"{result:.{places}f}"
    if places:
        rendered = rendered.rstrip("0").rstrip(".")
    return rendered or "0"


def yes_no(value: Any) -> bool | None:
    text = normalize(value)
    if text in {"si", "s", "yes", "1", "true"}:
        return True
    if text in {"no", "n", "0", "false"}:
        return False
    return None


def split_names(value: Any, missing: str) -> tuple[str, str, bool]:
    parts = [part for part in re.split(r"\s+", ascii_text(value).upper()) if part]
    if not parts:
        return "", missing, False
    if len(parts) == 1:
        return parts[0], missing, False
    return parts[0], " ".join(parts[1:]), len(parts) > 2


def extract_dated_labs(
    text: Any,
    source: str,
    source_kind: str,
    source_rank: int,
    encounter_date: date | None = None,
) -> list[dict[str, Any]]:
    content = str(text or "")
    date_matches = list(NARRATIVE_DATE_PATTERN.finditer(content))
    labs: list[dict[str, Any]] = []
    seen: set[tuple[str, date, Decimal]] = set()
    for index, date_match in enumerate(date_matches):
        year = int(date_match.group("year"))
        year += 2000 if year < 50 else 1900 if year < 100 else 0
        try:
            lab_date = date(
                year,
                int(date_match.group("month")),
                int(date_match.group("day")),
            )
        except ValueError:
            continue
        segment_end = (
            date_matches[index + 1].start()
            if index + 1 < len(date_matches)
            else len(content)
        )
        segment = content[date_match.end() : segment_end]
        for lab_type, pattern in NARRATIVE_LAB_PATTERNS.items():
            lower, upper = NARRATIVE_LAB_RANGES[lab_type]
            for lab_match in pattern.finditer(segment):
                value = decimal_value(lab_match.group(1))
                if value is None or not lower <= value <= upper:
                    continue
                identity = (lab_type, lab_date, value)
                if identity in seen:
                    continue
                seen.add(identity)
                labs.append(
                    {
                        "type": lab_type,
                        "date": lab_date,
                        "value": value,
                        "source": source,
                        "source_kind": source_kind,
                        "source_rank": source_rank,
                        "encounter_date": encounter_date,
                    }
                )
    return labs


def lab_sort_key(item: dict[str, Any]) -> tuple[bool, date, int, date]:
    lab_date = item.get("date")
    return (
        lab_date is not None,
        lab_date or date.min,
        int(item.get("source_rank") or 0),
        item.get("encounter_date") or date.min,
    )


def normalized_vital_value(vital_type: str, value: Any) -> Decimal | None:
    number = decimal_value(value)
    if number is None:
        return None
    if vital_type == "height" and Decimal("0.3") <= number < Decimal("3"):
        number *= Decimal(100)
    rule = VITAL_FIELDS[vital_type]
    if not rule["minimum"] <= number <= rule["maximum"]:
        return None
    return number


def extract_support_vitals(
    layout_pages: list[str],
    source: str,
    measurement_date: date | None,
) -> list[dict[str, Any]]:
    selected: dict[str, dict[str, Any]] = {}
    pending_by_column: dict[int, str | None] = {}
    in_vital_section = False
    number_pattern = re.compile(r"(?<![\d.])-?\d+(?:[.,]\d+)?")
    table_labels = [
        (vital_type, rule["label"])
        for vital_type, rule in VITAL_FIELDS.items()
    ] + [
        (None, re.compile(pattern, re.I))
        for pattern in (
            r"Frecuencia\s+respiratoria",
            r"Saturaci[oó]n\s+de\s+ox[ií]geno",
            r"Temperatura",
            r"Frecuencia\s+cardiaca",
            r"Per[ií]metro\s+abdominal",
            r"\bIMC\b",
        )
    ]
    ignored_lines = (
        "neurum sas",
        "nit:",
        "punto clave",
        "servicioalcliente",
        "tel:",
        "documento:",
        "fecha:",
        "pagina ",
    )

    def column_for(position: int) -> int:
        if position < 38:
            return 0
        if position < 82:
            return 1
        return 2

    for page_text in layout_pages:
        for line in page_text.splitlines():
            normalized_line = normalize(line)
            if "signos vitales" in normalized_line:
                in_vital_section = True
                pending_by_column = {}
                continue
            if not in_vital_section:
                continue
            if "examen fisico" in normalized_line or "control metabolico" in normalized_line:
                in_vital_section = False
                pending_by_column = {}
                continue

            if not normalized_line:
                continue
            if any(marker in normalized_line for marker in ignored_lines):
                continue

            # En algunos PDF el valor de una columna aparece en la misma línea
            # donde inicia la etiqueta de otra columna. Se consumen primero los
            # valores pendientes y después se actualizan las etiquetas.
            for candidate in number_pattern.finditer(line):
                column = column_for(candidate.start())
                vital_type = pending_by_column.get(column)
                if not vital_type:
                    continue
                value = normalized_vital_value(vital_type, candidate.group(0))
                if value is None:
                    continue
                selected[vital_type] = {
                    "type": vital_type,
                    "value": value,
                    "date": measurement_date,
                    "source": source,
                    "source_kind": "pdf_vital",
                    "source_rank": 2,
                }
                pending_by_column[column] = None

            labels_by_column: dict[int, tuple[int, str | None]] = {}
            for vital_type, pattern in table_labels:
                match = pattern.search(line)
                if not match:
                    continue
                column = column_for(match.start())
                current = labels_by_column.get(column)
                if current is None or match.start() < current[0]:
                    labels_by_column[column] = (match.start(), vital_type)
            for column, (_, vital_type) in labels_by_column.items():
                pending_by_column[column] = vital_type
    return list(selected.values())


def extract_support(path: Path) -> SupportEvidence:
    display_name = path.name.split("__", 1)[-1]
    match = re.match(
        r"^(?P<type>[A-Za-z]+)(?P<doc>\d+)_(?P<specialty>.+)_(?P<date>\d{4}-\d{2}-\d{2})\.pdf$",
        display_name,
        re.I,
    )
    doc_type = match.group("type").upper() if match else ""
    document = match.group("doc") if match else ""
    specialty = match.group("specialty").replace("_", " ") if match else ""
    attention_date = parse_date(match.group("date")) if match else None
    page_texts = []
    layout_pages = []
    with path.open("rb") as pdf_stream:
        reader = PdfReader(pdf_stream)
        for page in reader.pages:
            extracted = page.extract_text() or ""
            page_texts.append(extracted)
            try:
                layout_pages.append(
                    page.extract_text(extraction_mode="layout") or extracted
                )
            except (KeyError, TypeError):
                layout_pages.append(extracted)
    text = "\n".join(page_texts)
    compact = re.sub(r"\s+", " ", text)
    normalized = normalize(compact)
    phones = re.findall(r"(?<!\d)(3\d{9})(?!\d)", compact)
    ips_codes = re.findall(r"REPS\s*0*(\d{11,12})", compact, re.I)
    lab_pattern = re.compile(
        r"(Creatinina|HbA1c|Albuminuria|RAC|Colesterol LDL|Colesterol HDL|Colesterol Total)"
        r"\s+(\d{2}-\d{2}-\d{4}|-)\s+([0-9]+(?:\.[0-9]+)?|-)"
        r"\s+(\d{2}-\d{2}-\d{4}|-)\s+([0-9]+(?:\.[0-9]+)?|-)",
        re.I,
    )
    labs = []
    for lab_match in lab_pattern.finditer(compact):
        label = normalize(lab_match.group(1))
        latest_date = parse_date(lab_match.group(4))
        latest_value = decimal_value(lab_match.group(5))
        if latest_date and latest_value is not None:
            labs.append(
                {
                    "type": LAB_LABELS[label],
                    "date": latest_date,
                    "value": latest_value,
                    "source": display_name,
                    "source_kind": "pdf_structured",
                    "source_rank": 4,
                    "encounter_date": attention_date,
                }
            )
    narrative_labs = extract_dated_labs(
        compact,
        display_name,
        "pdf_narrative",
        3,
        attention_date,
    )
    known_labs = {(lab["type"], lab["date"], lab["value"]) for lab in labs}
    labs.extend(
        lab
        for lab in narrative_labs
        if (lab["type"], lab["date"], lab["value"]) not in known_labs
    )
    vitals = extract_support_vitals(layout_pages, display_name, attention_date)
    return SupportEvidence(
        file_name=display_name,
        document=document,
        doc_type=doc_type,
        attention_date=attention_date,
        specialty=specialty,
        text=text,
        normalized_text=normalized,
        phone=phones[0] if phones else "",
        ips_code=ips_codes[-1].zfill(12) if ips_codes else "",
        labs=labs,
        vitals=vitals,
    )


def load_consultations(path: Path) -> tuple[dict[str, list[dict[str, Any]]], list[str]]:
    workbook = load_workbook(path, read_only=True, data_only=True)
    sheet = workbook.active
    rows = sheet.iter_rows(values_only=True)
    try:
        headers = [str(value or "").strip() for value in next(rows)]
    except StopIteration as exc:
        raise ValueError("El archivo de atenciones no contiene filas.") from exc
    required = {"numero_documento", "fechaCreacion", "nombres_paciente", "apellidos_paciente"}
    missing = sorted(required.difference(headers))
    if missing:
        raise ValueError("Faltan columnas obligatorias en atenciones: " + ", ".join(missing))
    grouped: dict[str, list[dict[str, Any]]] = defaultdict(list)
    for values in rows:
        record = dict(zip(headers, values))
        document = clean_document(record.get("numero_documento"))
        if document:
            record["_encounter_date"] = parse_date(record.get("fechaCreacion"))
            record["_narrative_labs"] = extract_dated_labs(
                record.get("paraclinicos"),
                "Narrativa de atención",
                "excel_narrative",
                2,
                record["_encounter_date"],
            )
            grouped[document].append(record)
    for records in grouped.values():
        records.sort(key=lambda item: item.get("_encounter_date") or date.min)
    return grouped, headers


def consolidate(records: list[dict[str, Any]]) -> dict[str, Any]:
    merged: dict[str, Any] = {}
    for record in records:
        for key, value in record.items():
            if meaningful(value):
                merged[key] = value
    merged["_encounter_dates"] = [
        record["_encounter_date"] for record in records if record.get("_encounter_date")
    ]
    return merged


def latest_excel_lab(
    records: list[dict[str, Any]],
    field: str,
    lab_type: str,
) -> dict[str, Any] | None:
    candidates = []
    for record in records:
        value = decimal_value(record.get(field))
        lab_date = parse_date(record.get("fecha_paraclinicos"))
        encounter_date = record.get("_encounter_date")
        if value is not None:
            candidates.append(
                {
                    "type": lab_type,
                    "date": lab_date,
                    "value": value,
                    "source": "Dato discreto de atención",
                    "source_kind": "excel_discrete",
                    "source_rank": 1,
                    "encounter_date": encounter_date,
                }
            )
        narrative_labs = record.get("_narrative_labs")
        if narrative_labs is None:
            narrative_labs = extract_dated_labs(
                record.get("paraclinicos"),
                "Narrativa de atención",
                "excel_narrative",
                2,
                encounter_date,
            )
        candidates.extend(lab for lab in narrative_labs if lab["type"] == lab_type)
    if not candidates:
        return None
    return max(candidates, key=lab_sort_key)


def choose_lab(
    records: list[dict[str, Any]],
    supports: list[SupportEvidence],
    excel_field: str,
    pdf_type: str,
    warnings: list[str],
) -> dict[str, Any] | None:
    candidates = []
    excel = latest_excel_lab(records, excel_field, pdf_type)
    if excel:
        candidates.append(excel)
    for support in supports:
        candidates.extend(lab for lab in support.labs if lab["type"] == pdf_type)
    if not candidates:
        return None
    selected = max(candidates, key=lab_sort_key)
    if not selected.get("date"):
        warnings.append(f"{pdf_type}: valor encontrado sin fecha propia de toma")
    elif selected.get("source_kind") == "excel_narrative":
        warnings.append(f"{pdf_type}: fecha y valor recuperados de la narrativa de atención")
    elif selected.get("source_kind") == "pdf_narrative":
        warnings.append(f"{pdf_type}: fecha y valor recuperados de la narrativa del PDF")
    if excel and selected.get("source_kind", "").startswith("pdf"):
        if selected["date"] and (not excel["date"] or selected["date"] >= excel["date"]):
            if selected["value"] != excel["value"] or selected["date"] != excel["date"]:
                warnings.append(f"{pdf_type}: el PDF actualiza el valor de la descarga")
    return selected


def vital_sort_key(item: dict[str, Any]) -> tuple[bool, date, int]:
    measurement_date = item.get("date")
    return (
        measurement_date is not None,
        measurement_date or date.min,
        int(item.get("source_rank") or 0),
    )


def choose_vitals(
    records: list[dict[str, Any]],
    supports: list[SupportEvidence],
) -> dict[str, dict[str, Any] | None]:
    candidates: dict[str, list[dict[str, Any]]] = defaultdict(list)
    for record in records:
        encounter_date = record.get("_encounter_date")
        for vital_type, rule in VITAL_FIELDS.items():
            value = normalized_vital_value(
                vital_type,
                record.get(rule["attention_field"]),
            )
            if value is None:
                continue
            candidates[vital_type].append(
                {
                    "type": vital_type,
                    "value": value,
                    "date": encounter_date,
                    "source": "Atención clínica",
                    "source_kind": "attention_vital",
                    "source_rank": 1,
                }
            )
    for support in supports:
        for vital in support.vitals:
            if vital.get("type") in VITAL_FIELDS:
                candidates[vital["type"]].append(vital)
    return {
        vital_type: (
            max(candidates[vital_type], key=vital_sort_key)
            if candidates[vital_type]
            else None
        )
        for vital_type in VITAL_FIELDS
    }


def build_patient_row(
    document: str,
    records: list[dict[str, Any]],
    supports: list[SupportEvidence],
    settings: dict[str, Any],
    headers: list[str],
    divipola: dict[str, str],
) -> tuple[dict[str, Any], dict[str, Any]]:
    merged = consolidate(records)
    values = {header: "" for header in headers}
    errors: list[str] = []
    warnings: list[str] = []
    evidence: list[str] = []

    first_name, second_name, complex_names = split_names(merged.get("nombres_paciente"), "NONE")
    first_surname, second_surname, complex_surnames = split_names(merged.get("apellidos_paciente"), "NOAP")
    if complex_names or complex_surnames:
        warnings.append("Verificar separación de nombres y apellidos")

    source_doc_type = next(
        (
            code
            for record in reversed(records)
            if (
                code := DOCUMENT_TYPES.get(
                    normalize_key(record.get("tipo_documento")),
                    "",
                )
            )
        ),
        "",
    )
    support_types = sorted({support.doc_type for support in supports if support.doc_type})
    if not source_doc_type and len(support_types) == 1:
        source_doc_type = support_types[0]
        warnings.append("Tipo documental completado desde el nombre del soporte PDF")
    if support_types and source_doc_type and source_doc_type not in support_types:
        errors.append(
            f"Tipo documental no coincide: Excel {source_doc_type}, PDF {'/'.join(support_types)}"
        )

    sex = SEX_CODES.get(normalize_key(merged.get("sexo")), "")
    regimen = REGIMEN_CODES.get(normalize_key(merged.get("regimen")), "")
    municipality_name = normalize_key(merged.get("municipio_residencia"))
    municipality_code = divipola.get(municipality_name, "")
    latest_support = max(
        supports,
        key=lambda item: item.attention_date or date.min,
        default=None,
    )
    phone = next((support.phone for support in reversed(supports) if support.phone), "") or "0"
    ips_code = str(settings.get("ips_code") or "").strip()
    if not ips_code and latest_support:
        ips_code = latest_support.ips_code

    values.update(
        {
            "VAR1_PrimerNombre": first_name,
            "VAR2_SegundoNombre": second_name,
            "VAR3_PrimerApellido": first_surname,
            "VAR4_SegundoApellido": second_surname,
            "VAR5_TipoIdentificacion": source_doc_type,
            "VAR6_Identificacion": document,
            "VAR7_FechaNacimiento": date_text(merged.get("fecha_nacimiento")),
            "VAR8_Sexo": sex,
            "VAR9_idRegimen": regimen,
            "VAR10_idEPS": str(settings.get("eapb_code") or "").strip(),
            "VAR11_idPertenenciaEtnica": str(settings.get("ethnicity_code") or "6"),
            "VAR12_idGrupoPoblacional": str(settings.get("population_group") or "61"),
            "VAR13_idMunicipio": municipality_code,
            "VAR14_TelefonoPaciente": phone,
            "VAR15_FechaAfiliacion": date_text(settings.get("affiliation_date")),
            "VAR16_CodigoIPS": ips_code,
            "VAR17_FechaIngProgNefro": date_text(
                settings.get("program_entry_date"), "1800-01-01"
            ),
        }
    )

    hta = yes_no(merged.get("hta"))
    values["VAR18_DiagnosticoHTA"] = 1 if hta else 2
    values["VAR19_FechaDxHTA"] = (
        date_text(merged.get("fechaDiagnosticoHTA"), "1800-01-01")
        if hta
        else "1845-01-01"
    )
    values["VAR19_1_CostoHTA"] = (
        merged.get("costo_hta")
        if meaningful(merged.get("costo_hta"))
        else settings.get("htn_cost_default", "")
    )

    dm_code = DM_CODES.get(normalize_key(merged.get("tipo_diabetes")), 4)
    values["VAR20_DiagnosticoDM"] = dm_code
    values["VAR21_FechaDxDM"] = date_text(
        merged.get("fecha_diagnostico_diabetes"), "1800-01-01"
    )
    values["VAR21_1_CostoDM"] = (
        merged.get("costo_dm")
        if meaningful(merged.get("costo_dm"))
        else settings.get("dm_cost_default", "")
    )

    vitals = choose_vitals(records, supports)
    weight_vital = vitals["weight"]
    height_vital = vitals["height"]
    systolic_vital = vitals["systolic"]
    diastolic_vital = vitals["diastolic"]
    values["VAR23_Peso"] = (
        truncated(weight_vital["value"]) if weight_vital else ""
    )
    values["VAR24_Talla"] = (
        truncated(height_vital["value"]) if height_vital else ""
    )
    values["VAR25_Tas"] = (
        truncated(systolic_vital["value"], 0) if systolic_vital else 999
    )
    values["VAR26_Tad"] = (
        truncated(diastolic_vital["value"], 0) if diastolic_vital else 999
    )

    creatinine = choose_lab(records, supports, "creatinine", "creatinine", warnings)
    hba1c = choose_lab(records, supports, "hba1c", "hba1c", warnings)
    rac = choose_lab(records, supports, "rac", "rac", warnings)
    ldl = choose_lab(records, supports, "_no_ldl_column", "ldl", warnings)
    hdl = choose_lab(records, supports, "_no_hdl_column", "hdl", warnings)
    total_cholesterol = choose_lab(
        records, supports, "_no_total_column", "total_cholesterol", warnings
    )

    values["VAR27_Creatinina"] = truncated(creatinine["value"]) if creatinine else 99
    values["VAR27_1_FchUltimaCreatinina"] = (
        creatinine["date"].isoformat() if creatinine and creatinine["date"] else "1800-01-01"
    )
    values["VAR28_HemoglobinaGlicosilada"] = truncated(hba1c["value"]) if hba1c else 99
    values["VAR28_1_FchUltimaHemoglobinaGlicosilada"] = (
        hba1c["date"].isoformat() if hba1c and hba1c["date"] else "1800-01-01"
    )
    # La descarga registra microalbuminuria en mg/L, no albuminuria de 24 horas.
    values["VAR29_Albuminuriamg24h"] = 9999
    values["VAR29_1_FchUltimaAlbuminuria"] = "1800-01-01"
    values["VAR30_Creatinuria"] = truncated(rac["value"]) if rac else 9999
    values["VAR30_1_FchUltimaCreatinuria"] = (
        rac["date"].isoformat() if rac and rac["date"] else "1800-01-01"
    )
    for label, lab in (
        ("Creatinina", creatinine),
        ("HbA1c", hba1c),
        ("RAC", rac),
    ):
        if lab and not lab.get("date"):
            errors.append(f"{label} tiene valor, pero no fecha identificable de toma")
    for value_header, date_header, lab, label in (
        (
            "VAR31_Colesterol",
            "VAR31_1_FchUltimoColesterol",
            total_cholesterol,
            "Colesterol total",
        ),
        ("VAR32_HDL", "VAR32_1_FchUltimoHDL", hdl, "HDL"),
        ("VAR33_LDL", "VAR33_1_FchUltimoLDL", ldl, "LDL"),
    ):
        values[value_header] = truncated(lab["value"]) if lab else 999
        values[date_header] = (
            lab["date"].isoformat()
            if lab and lab["date"]
            else "1800-01-01"
            if lab
            else "1845-01-01"
        )
        if lab and not lab.get("date"):
            errors.append(f"{label} tiene valor, pero no fecha identificable de toma")

    raw_tfg = decimal_value(merged.get("tfg"))
    tfg = None
    if raw_tfg is None:
        values["VAR35_TasaTFG"] = 999
        errors.append(
            "TFG no registrada en la atención; no se calcula automáticamente"
        )
    elif raw_tfg in {Decimal(999), Decimal(988), Decimal(777)}:
        values["VAR35_TasaTFG"] = truncated(raw_tfg, 0)
        evidence.append(f"Código especial de TFG {truncated(raw_tfg, 0)} registrado en la atención")
    elif raw_tfg <= 0:
        values["VAR35_TasaTFG"] = 999
        errors.append("TFG registrada inválida; no se reemplaza mediante cálculo")
    else:
        tfg = raw_tfg
        values["VAR35_TasaTFG"] = truncated(raw_tfg)
        evidence.append("TFG tomada exactamente de la atención")
        birth_date = parse_date(merged.get("fecha_nacimiento"))
        reference_date = (
            creatinine["date"]
            if creatinine and creatinine["date"]
            else parse_date(merged.get("fechaCreacion"))
        )
        if birth_date and reference_date:
            age = reference_date.year - birth_date.year - (
                (reference_date.month, reference_date.day)
                < (birth_date.month, birth_date.day)
            )
            if age < 18:
                warnings.append(
                    "TFG pediátrica registrada: validar que corresponda a Schwartz"
                )
            else:
                warnings.append(
                    "TFG adulta registrada: validar que corresponda a Cockcroft-Gault"
                )

    support_text = " ".join(support.normalized_text for support in supports)
    values["VAR36_IECA"] = 1 if any(name in support_text for name in IECA_NAMES) else 2
    values["VAR37_ARAII"] = 1 if any(name in support_text for name in ARA2_NAMES) else 2
    if not supports:
        warnings.append("IECA/ARAII inferidos sin soporte PDF cargado")

    explicit_erc = yes_no(merged.get("nefropatia_diabetica"))
    abnormal_tfg = tfg is not None and tfg < 60
    abnormal_rac = rac is not None and rac["value"] >= 30
    has_renal_tests = creatinine is not None or rac is not None
    if explicit_erc is True:
        erc_status = 1
    elif abnormal_tfg or abnormal_rac:
        erc_status = 2
        warnings.append("Alteración renal sin evidencia suficiente de persistencia: indeterminado")
    elif has_renal_tests:
        erc_status = 0
    else:
        erc_status = 3
    if erc_status == 1:
        if tfg is None:
            stage = 99
        elif tfg >= 90:
            stage = 1
        elif tfg >= 60:
            stage = 2
        elif tfg >= 30:
            stage = 3
        elif tfg >= 15:
            stage = 4
        else:
            stage = 5
    else:
        stage = 98 if erc_status == 0 else 99
    if values["VAR35_TasaTFG"] == 999:
        stage = 99
    values["VAR22_EtiologiaERC"] = 7 if erc_status == 1 else 98
    values["VAR34_PTH"] = 9999 if erc_status == 1 else 9988
    values["VAR34_1_FchUltimoPTH"] = "1800-01-01" if erc_status == 1 else "1845-01-01"
    values["VAR38_DiagnosticoERC"] = erc_status
    values["VAR39_EstadioERC"] = stage
    values["VAR40_FechaDiagnosticoERC"] = (
        date_text(merged.get("fecha_diagnostico_nefropatia"), "1800-01-01")
        if stage == 5
        else "1845-01-01"
    )
    values["VAR41_ProgramaAtencionERC"] = 99

    if creatinine and creatinine["date"]:
        if stage in {3, 4}:
            creatinine_start = date(2026, 1, 1)
        elif stage == 5:
            creatinine_start = date(2026, 4, 1)
        else:
            creatinine_start = REPORT_START
        if not creatinine_start <= creatinine["date"] <= REPORT_END:
            errors.append("Creatinina fuera de la ventana exigida para el estadio")
        if (
            weight_vital
            and weight_vital.get("date")
            and abs((creatinine["date"] - weight_vital["date"]).days) > 183
        ):
            errors.append("Peso y creatinina tienen un intervalo mayor a 6 meses")
    if hba1c and hba1c["date"] and not date(2026, 1, 1) <= hba1c["date"] <= REPORT_END:
        errors.append("HbA1c fuera del último semestre del reporte")
    if rac and rac["date"] and not REPORT_START <= rac["date"] <= REPORT_END:
        errors.append("RAC fuera del periodo de reporte")
    for label, lab in (
        ("Colesterol total", total_cholesterol),
        ("HDL", hdl),
        ("LDL", ldl),
    ):
        if lab and lab["date"] and not REPORT_START <= lab["date"] <= REPORT_END:
            errors.append(f"{label} fuera del periodo de reporte")

    transplant = yes_no(merged.get("transplant")) is True
    trr = yes_no(merged.get("trr")) is True or transplant
    if trr:
        errors.append("Paciente con TRR/trasplante: requiere completar variables 42 a 75 manualmente")
    else:
        values.update(
            {
                "VAR42_TasaTFGenTRR": 98,
                "VAR43_ModoInicioTRR": 97,
                "VAR44_FechaTRR": "1845-01-01",
                "VAR45_FechIngresoUnidRenalActual": "1845-01-01",
                "VAR46_TerapiaHD": 98,
                "VAR47_KTVsp": 98,
                "VAR48_CostoHD": 98,
                "VAR49_TerapiaDP": 98,
                "VAR50_KTVdpd": 98,
                "VAR51_HorasDialisis": 98,
                "VAR52_Peritonitis": 98,
                "VAR53_CostoDP": 98,
                "VAR54_VacHepatitisB": 98,
                "VAR55_FchDxHepatitisB": "1811-01-01",
                "VAR56_FchDxHepatitisC": "1811-01-01",
                "VAR57_TerapiaERC": 2,
                "VAR58_CostoERC": 98,
                "VAR59_Hemoglobina": 99 if erc_status == 1 else 98,
                "VAR60_AlbuminaSerica": 99 if erc_status == 1 else 98,
                "VAR61_Fosforo": 99 if erc_status == 1 else 98,
            }
        )
        transplant_review = 98 if erc_status == 0 else 97 if stage in {1, 2, 3, 4} else 99
        for header in (
            "VAR62_ValInicialNefro",
            "VAR62_1_CntrIndicTpltCancer",
            "VAR62_2_CntrIndicTpltInfeccionNoTratada",
            "VAR62_3_CntrIndicTpltPacienteNoAutoriza",
            "VAR62_4_CntrIndicTpltEsperanzaVida",
            "VAR62_5_CntrIndicTpltLimitacionesAutocuidado",
            "VAR62_6_CntrIndicTpltEnfCardCerebvascVascPerif",
            "VAR62_7_CntrIndicTpltInfeccionVIH",
            "VAR62_8_CntrIndicTpltInfeccionVCH",
            "VAR62_9_CntrIndicTpltEnfermedadInmunologica",
            "VAR62_10_CntrIndicTpltEnfermedadPulmonarCronica",
            "VAR62_11_CntrIndicTpltOtrasEnfermedadesCronicas",
        ):
            values[header] = transplant_review
        values.update(
            {
                "VAR63_FchIngresoListaEsperaTpl": "1849-01-01",
                "VAR63_1_idIPSListaEspera": 90,
                "VAR64_TrasplanteRenal": 5,
                "VAR65_idEpsTransplante": 98,
                "VAR66_idGrupoTransplante": 98,
                "VAR67_idTipoDonante": 98,
                "VAR68_CostoTrasplante": 98,
                "VAR69_idComplicacionTrasplante": 98,
            }
        )
        for header in (
            "VAR69_1_FchDxInfeccionCitomegalovirus",
            "VAR69_2_FchDxInfeccionHongos",
            "VAR69_3_FchDxInfeccionTb",
            "VAR69_4_FchDxComplicVascular",
            "VAR69_5_FchDxComplicUrologica",
            "VAR69_6_FchDxComplicHeridaQx",
            "VAR69_7_Fch1erDxCancer",
        ):
            values[header] = "1845-01-01"
        for header in (
            "VAR70_CuantosMedicInmunosupresores",
            "VAR70_1_RecibioMetilprednisolona",
            "VAR70_2_RecibioAzatioprina",
            "VAR70_3_RecibioCiclosporina",
            "VAR70_4_RecibioMicofenolato",
            "VAR70_5_RecibioTacrolimus",
            "VAR70_6_RecibioPrednisona",
            "VAR70_7_RecibioMedicamento01",
            "VAR70_8_RecibioMedicamento02",
            "VAR70_9_RecibioMedicamento03",
            "VAR71_EpisodiosRechazo",
            "VAR74_CantTrasplantes",
            "VAR75_CostoTerapiaPostransplante",
        ):
            values[header] = 98
        values["VAR72_Fch1erRechazoagudoInjerto"] = "1845-01-01"
        values["VAR73_FchRetornoDialisis"] = "1845-01-01"

    service_months = {
        (encounter.year, encounter.month)
        for encounter in merged["_encounter_dates"]
        if REPORT_START <= encounter <= REPORT_END
    }
    values["VAR76_TiempoPrestacionServicios"] = len(service_months)
    values["VAR77_CostoTotal"] = (
        merged.get("costo_total")
        if meaningful(merged.get("costo_total"))
        else settings.get("total_cost_default", "")
    )
    values["VAR78_idEpsOrigen"] = 98
    values["VAR79_Novedad"] = 98
    values["VAR80_idCausaMuerte"] = 98
    values["VAR80_1_FchMuerte"] = "1845-01-01"
    values["VAR81_SerialBDUA"] = (
        merged.get("serial_bdua")
        if meaningful(merged.get("serial_bdua"))
        else (0 if settings.get("territorial_entity") else "")
    )
    values["VAR82_V134FechadeCorte"] = date_text(
        settings.get("cutoff_date"), REPORT_END.isoformat()
    )

    required_missing = [header for header in headers if not meaningful(values.get(header))]
    friendly_required = {
        "VAR10_idEPS": "Código EAPB",
        "VAR13_idMunicipio": "DIVIPOLA municipio",
        "VAR15_FechaAfiliacion": "Fecha de afiliación",
        "VAR16_CodigoIPS": "Código IPS",
        "VAR19_1_CostoHTA": "Costo HTA",
        "VAR21_1_CostoDM": "Costo DM",
        "VAR23_Peso": "Peso",
        "VAR24_Talla": "Talla",
        "VAR77_CostoTotal": "Costo total",
        "VAR81_SerialBDUA": "Serial BDUA",
    }
    for header in required_missing:
        errors.append(f"Falta {friendly_required.get(header, header)}")

    support_count = len(supports)
    if not support_count:
        warnings.append("Sin soporte PDF coincidente")
    else:
        evidence.append(f"{support_count} soporte(s) PDF asociado(s)")
    status = "LISTO" if not errors else "SIN SOPORTE" if not supports else "REVISAR"
    values["Auditoria"] = status
    coverage = round(
        100
        * sum(1 for header in headers if meaningful(values.get(header)))
        / max(len(headers), 1),
        1,
    )
    review = {
        "document": document,
        "document_type": source_doc_type,
        "patient": " ".join(
            part
            for part in (
                first_name,
                second_name if second_name != "NONE" else "",
                first_surname,
                second_surname if second_surname != "NOAP" else "",
            )
            if part
        ),
        "status": status,
        "coverage": coverage,
        "support_count": support_count,
        "latest_support": latest_support.attention_date.isoformat()
        if latest_support and latest_support.attention_date
        else "",
        "errors": errors,
        "warnings": warnings,
        "evidence": evidence,
        "key_values": {
            "weight": values["VAR23_Peso"],
            "weight_date": (
                weight_vital["date"].isoformat()
                if weight_vital and weight_vital.get("date")
                else ""
            ),
            "weight_source": weight_vital["source"] if weight_vital else "",
            "height": values["VAR24_Talla"],
            "height_date": (
                height_vital["date"].isoformat()
                if height_vital and height_vital.get("date")
                else ""
            ),
            "height_source": height_vital["source"] if height_vital else "",
            "systolic": values["VAR25_Tas"],
            "systolic_date": (
                systolic_vital["date"].isoformat()
                if systolic_vital and systolic_vital.get("date")
                else ""
            ),
            "systolic_source": systolic_vital["source"] if systolic_vital else "",
            "diastolic": values["VAR26_Tad"],
            "diastolic_date": (
                diastolic_vital["date"].isoformat()
                if diastolic_vital and diastolic_vital.get("date")
                else ""
            ),
            "diastolic_source": diastolic_vital["source"] if diastolic_vital else "",
            "creatinine": values["VAR27_Creatinina"],
            "creatinine_date": values["VAR27_1_FchUltimaCreatinina"],
            "creatinine_source": creatinine["source"] if creatinine else "",
            "hba1c": values["VAR28_HemoglobinaGlicosilada"],
            "hba1c_date": values["VAR28_1_FchUltimaHemoglobinaGlicosilada"],
            "hba1c_source": hba1c["source"] if hba1c else "",
            "rac": values["VAR30_Creatinuria"],
            "rac_date": values["VAR30_1_FchUltimaCreatinuria"],
            "rac_source": rac["source"] if rac else "",
            "tfg": values["VAR35_TasaTFG"],
            "erc_status": values["VAR38_DiagnosticoERC"],
            "erc_stage": values["VAR39_EstadioERC"],
        },
    }
    return values, review


def validate_settings(settings: dict[str, Any]) -> dict[str, Any]:
    cleaned = {
        "cutoff_date": date_text(settings.get("cutoff_date"), REPORT_END.isoformat()),
        "eapb_code": ascii_text(settings.get("eapb_code")).upper(),
        "ips_code": re.sub(r"\D", "", str(settings.get("ips_code") or "")),
        "ethnicity_code": str(settings.get("ethnicity_code") or "6").strip(),
        "population_group": str(settings.get("population_group") or "61").strip(),
        "affiliation_date": date_text(settings.get("affiliation_date")),
        "program_entry_date": date_text(settings.get("program_entry_date")),
        "territorial_entity": bool(settings.get("territorial_entity")),
        "htn_cost_default": truncated(settings.get("htn_cost_default")),
        "dm_cost_default": truncated(settings.get("dm_cost_default")),
        "total_cost_default": truncated(settings.get("total_cost_default")),
    }
    if cleaned["ips_code"] and len(cleaned["ips_code"]) != 12:
        raise ValueError("El código REPS de la IPS debe tener 12 dígitos.")
    cutoff = parse_date(cleaned["cutoff_date"])
    if not cutoff:
        raise ValueError("La fecha de corte no es válida.")
    return cleaned


def write_outputs(
    output_xlsx: Path,
    output_txt: Path,
    rows: list[dict[str, Any]],
    reviews: list[dict[str, Any]],
    headers: list[str],
    settings: dict[str, Any],
) -> int:
    workbook = load_workbook(TEMPLATE_PATH)
    sheet = workbook.active
    sheet.title = "Malla_CAC"
    sheet.freeze_panes = "A3"
    for row_number, values in enumerate(rows, start=3):
        for column_number, header in enumerate(["Auditoria", *headers], start=1):
            cell = sheet.cell(row_number, column_number, values.get(header, ""))
            cell.alignment = Alignment(vertical="top")
        audit_cell = sheet.cell(row_number, 1)
        if audit_cell.value == "LISTO":
            audit_cell.fill = PatternFill("solid", fgColor="D9EAD3")
        elif audit_cell.value == "SIN SOPORTE":
            audit_cell.fill = PatternFill("solid", fgColor="FCE5CD")
        else:
            audit_cell.fill = PatternFill("solid", fgColor="F4CCCC")
        audit_cell.font = Font(bold=True)
    sheet.auto_filter.ref = f"A2:{get_column_letter(len(headers) + 1)}{max(3, len(rows) + 2)}"

    validation = workbook.create_sheet("Validacion")
    validation_headers = [
        "Estado",
        "Tipo documento",
        "Documento",
        "Paciente",
        "Cobertura %",
        "Soportes PDF",
        "Último soporte",
        "TFG",
        "Peso",
        "Fecha peso",
        "Fuente peso",
        "Talla",
        "Fecha talla",
        "Fuente talla",
        "TAS",
        "Fecha TAS",
        "Fuente TAS",
        "TAD",
        "Fecha TAD",
        "Fuente TAD",
        "Creatinina",
        "Fecha creatinina",
        "Fuente creatinina",
        "HbA1c",
        "Fecha HbA1c",
        "Fuente HbA1c",
        "RAC",
        "Fecha RAC",
        "Fuente RAC",
        "Errores que bloquean",
        "Advertencias",
        "Evidencia",
    ]
    validation.append(validation_headers)
    for review, values in zip(reviews, rows):
        key = review["key_values"]
        validation.append(
            [
                review["status"],
                values.get("VAR5_TipoIdentificacion"),
                review["document"],
                review["patient"],
                review["coverage"],
                review["support_count"],
                review["latest_support"],
                key["tfg"],
                key["weight"],
                key["weight_date"],
                key["weight_source"],
                key["height"],
                key["height_date"],
                key["height_source"],
                key["systolic"],
                key["systolic_date"],
                key["systolic_source"],
                key["diastolic"],
                key["diastolic_date"],
                key["diastolic_source"],
                key["creatinine"],
                key["creatinine_date"],
                key["creatinine_source"],
                key["hba1c"],
                key["hba1c_date"],
                key["hba1c_source"],
                key["rac"],
                key["rac_date"],
                key["rac_source"],
                " | ".join(review["errors"]),
                " | ".join(review["warnings"]),
                " | ".join(review["evidence"]),
            ]
        )
    validation.freeze_panes = "A2"
    validation.auto_filter.ref = f"A1:AF{max(2, len(reviews) + 1)}"
    for cell in validation[1]:
        cell.fill = PatternFill("solid", fgColor="7F2A8A")
        cell.font = Font(color="FFFFFF", bold=True)
        cell.alignment = Alignment(wrap_text=True, vertical="center")
    widths = [
        14, 15, 18, 32, 12, 13, 16, 10,
        10, 16, 28, 10, 16, 28, 10, 16, 28, 10, 16, 28,
        12, 16, 28, 10, 16, 28, 10, 16, 28,
        62, 62, 42,
    ]
    for index, width in enumerate(widths, start=1):
        validation.column_dimensions[get_column_letter(index)].width = width
    for row in validation.iter_rows(min_row=2):
        row[29].alignment = Alignment(wrap_text=True, vertical="top")
        row[30].alignment = Alignment(wrap_text=True, vertical="top")
        row[31].alignment = Alignment(wrap_text=True, vertical="top")

    metadata = workbook.create_sheet("Configuracion")
    metadata.append(["Parámetro", "Valor", "Nota"])
    metadata_rows = [
        ("Fecha de corte", settings["cutoff_date"], "Variable 82"),
        ("Código EAPB", settings["eapb_code"], "Variable 10"),
        ("Código IPS", settings["ips_code"], "Variable 16"),
        ("Etnia", settings["ethnicity_code"], "Variable 11"),
        ("Grupo poblacional", settings["population_group"], "Variable 12"),
        (
            "Registros listos",
            sum(review["status"] == "LISTO" for review in reviews),
            "Sin errores bloqueantes",
        ),
        (
            "Registros por revisar",
            sum(review["status"] != "LISTO" for review in reviews),
            "No deben cargarse a SISCAC hasta corregirlos",
        ),
    ]
    for row in metadata_rows:
        metadata.append(row)
    for cell in metadata[1]:
        cell.fill = PatternFill("solid", fgColor="7F2A8A")
        cell.font = Font(color="FFFFFF", bold=True)
    metadata.column_dimensions["A"].width = 28
    metadata.column_dimensions["B"].width = 24
    metadata.column_dimensions["C"].width = 58

    output_xlsx.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(output_xlsx)

    ready_rows = [
        values
        for values, review in zip(rows, reviews)
        if review["status"] == "LISTO"
    ]
    if ready_rows:
        lines = []
        for values in ready_rows:
            fields = [ascii_text(values.get(header, "")) for header in headers]
            lines.append("\t".join(fields))
        output_txt.write_text("\r\n".join(lines) + "\r\n", encoding="cp1252")
    elif output_txt.exists():
        output_txt.unlink()
    return len(ready_rows)


def process_cac_job(
    consultations_path: Path,
    support_paths: list[Path],
    output_dir: Path,
    settings: dict[str, Any],
) -> dict[str, Any]:
    cleaned_settings = validate_settings(settings)
    grouped, source_headers = load_consultations(consultations_path)
    supports = [extract_support(path) for path in support_paths]
    supports_by_document: dict[str, list[SupportEvidence]] = defaultdict(list)
    for support in supports:
        if support.document:
            supports_by_document[support.document].append(support)
    for items in supports_by_document.values():
        items.sort(key=lambda item: item.attention_date or date.min)

    template = load_workbook(TEMPLATE_PATH, read_only=True, data_only=True)
    template_sheet = template.active
    output_headers = [
        str(template_sheet.cell(2, column).value or "").strip()
        for column in range(2, template_sheet.max_column + 1)
        if meaningful(template_sheet.cell(2, column).value)
    ]
    divipola = json.loads(DIVIPOLA_PATH.read_text(encoding="utf-8"))
    rows = []
    reviews = []
    for document in sorted(grouped, key=lambda value: (len(value), value)):
        row, review = build_patient_row(
            document,
            grouped[document],
            supports_by_document.get(document, []),
            cleaned_settings,
            output_headers,
            divipola,
        )
        rows.append(row)
        reviews.append(review)

    unmatched_supports = [
        support.file_name for support in supports if support.document not in grouped
    ]
    output_dir.mkdir(parents=True, exist_ok=True)
    output_xlsx = output_dir / "malla_cac_erc_validada.xlsx"
    eapb = cleaned_settings["eapb_code"] or "CODEAPB"
    cutoff_name = cleaned_settings["cutoff_date"].replace("-", "")
    output_txt = output_dir / f"{cutoff_name}_{eapb}_ERC.txt"
    ready = write_outputs(
        output_xlsx,
        output_txt,
        rows,
        reviews,
        output_headers,
        cleaned_settings,
    )

    error_counter = Counter(error for review in reviews for error in review["errors"])
    warning_counter = Counter(warning for review in reviews for warning in review["warnings"])
    summary = {
        "patients": len(rows),
        "source_records": sum(len(items) for items in grouped.values()),
        "source_columns": len(source_headers),
        "supports": len(supports),
        "matched_supports": len(supports) - len(unmatched_supports),
        "unmatched_supports": len(unmatched_supports),
        "unmatched_files": unmatched_supports[:100],
        "ready": ready,
        "review": len(rows) - ready,
        "without_support": sum(not review["support_count"] for review in reviews),
        "average_coverage": round(
            sum(review["coverage"] for review in reviews) / max(len(reviews), 1), 1
        ),
        "top_errors": error_counter.most_common(12),
        "top_warnings": warning_counter.most_common(12),
        "xlsx_name": output_xlsx.name,
        "txt_name": output_txt.name if ready else "",
        "generated_at": datetime.now().replace(microsecond=0).isoformat(),
    }
    preview = {
        "summary": summary,
        "settings": cleaned_settings,
        "patients": reviews,
    }
    (output_dir / "preview.json").write_text(
        json.dumps(preview, ensure_ascii=False, default=str),
        encoding="utf-8",
    )
    return preview
